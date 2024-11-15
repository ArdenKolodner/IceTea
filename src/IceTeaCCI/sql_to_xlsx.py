import argparse # Used for parsing command line arguments
import sys # Used for accessing command line arguments
import sqlite3 # Used for running the SQL script
import openpyxl # Used for creating the XLSX file
from openpyxl.comments import Comment # Used for saving metadata about each field
import json # Used for storing metadata in cell comments and file description, because it doesn't rely on whitespace like YAML
import os, platform, subprocess # Used for opening the generated XLSX file

def sql_to_xlsx(arglist=sys.argv[1:]):
  # Parse command line arguments
  parser = argparse.ArgumentParser()
  parser.add_argument("-i", "--input", help="Path to the SQL script to convert", default="./datasets/sample.sql")
  parser.add_argument("-o", "--output", help="Path to the XLSX file to create", default="./generated.xlsx")

  parser.add_argument("-ltn", "--log-table-names", help="Log names of converted tables", action="store_true")
  parser.add_argument("-lr", "--log-records", help="Log converted records", action="store_true")
  parser.add_argument("-lf", "--log-fields", help="Log names of detected fields in tables", action="store_true")

  output_options = parser.add_argument_group("Output options")
  output_options.add_argument("-of", "--open-file", help="Automatically open the generated XLSX file", action="store_true")
  output_options.add_argument("-nof", "--no-open-file", help="Do not open the generated XLSX file", action="store_false", dest='open_file')
  # By default, open the file at the end so user can edit data
  parser.set_defaults(open_file=True)

  args = parser.parse_args(arglist)

  # Output options
  LOG_TABLE_NAMES = args.log_table_names
  LOG_RECORDS = args.log_records
  LOG_FIELDS = args.log_fields

  OPEN_FILE = args.open_file

  INPUT_FILE = args.input
  OUTPUT_FILE = args.output

  AUTOGEN_PREFIX = "AUTOGENERATED, DO NOT EDIT!\n" # Must be kept identical with COMMENT_PREFIX in the other script

  # Logging functions
  def log_table_name(text):
    if LOG_TABLE_NAMES: print(text)
  def log_record(text):
    if LOG_RECORDS: print(text)
  def log_field(text):
    if LOG_FIELDS: print(text)

  with open(INPUT_FILE, 'r') as f:
    sql_script = f.read()

  # Create a database connection in memory, to hold the tables generated by the script
  conn = sqlite3.connect(":memory:")
  cur = conn.cursor()

  # Execute the script
  cur.executescript(sql_script)

  # Retrieve the names of the tables in the database
  tables = list(cur.execute("SELECT name FROM sqlite_master WHERE type='table'"))

  # Create a sample workbook
  workbook = openpyxl.Workbook()
  workbook.remove(workbook.active) # Remove the default worksheet

  # Keep track of the names of the worksheets, since they may be different from the table names.
  # This will be stored in the workbook description, so we can retrieve the original table names later.
  table_names = {}

  # Create a worksheet for each table
  for table in tables:
    table_name = table[0]

    # Excel limits sheet names to 31 characters. Thus we limit the sheet name to 29 characters, with the remaining 2 possibly added by OpenPyXL as a numeric suffix to avoid duplicates.
    # Currently it is not supported to have more than 100 fields that have the same first 29 characters. For now this is considered reasonable.
    sheet_target_name = table_name[:29] # If the table name is shorter than 29 chars, this just copies the whole thing.

    log_table_name(f"Creating worksheet for table: {table_name}")
    worksheet = workbook.create_sheet(sheet_target_name)
    sheet_name = worksheet.title # OpenPyXL may change the actual sheet name to avoid duplicates. Thus we must check what name was *actually* used.
    table_names[sheet_name] = table_name # Once we have a unique sheet name, we can link it to the (unique) table name.

    # Create header line
    fields = cur.execute(f"PRAGMA table_info('{table_name}')")
    for i, field in enumerate(fields):
      log_field(f"Field {i}: {field[1]}, type: {field[2]}, not null: {field[3]}, pk: {field[5]}")
      
      # Write header into table
      cell = worksheet.cell(row=1, column=i+1) # OpenPyXl uses 1-based indexing!
      cell.number_format = '@' # All field names are strings
      cell.value = field[1]

      # Prepare header metadata
      metadata = {"type": field[2], "not_null": bool(field[3]), "pk": int(field[5])}
      metadata_string = AUTOGEN_PREFIX + json.dumps(metadata)
      cell.comment = Comment(metadata_string, "AUTOGENERATED, DO NOT EDIT!")

    cur.execute(f"SELECT * FROM '{table_name}'")
    for i, row in enumerate(cur.fetchall()):
      for j, col in enumerate(row):
        cell = worksheet.cell(row=i+2, column=j+1) # OpenPyXl uses 1-based indexing!
        cell.value = col

      log_record(f"Table {table_name} ({sheet_name}): {col}")

  log_table_name(f"Table name links: {table_names}")

  workbook_metadata = {
    "table_names": table_names
  }
  workbook.properties.description = AUTOGEN_PREFIX + json.dumps(workbook_metadata)

  workbook.properties.creator = "IceTea"

  # Save file
  workbook.save(OUTPUT_FILE)

  # Open the created workbook
  if OPEN_FILE:
    if platform.system() == "Windows":
      os.startfile(OUTPUT_FILE)
    else:
      # subprocess.call() is safer than os.system(), as it prevents injection attacks. Of course, if an attacker can run this script, they probably have command-line access already, but why take the risk?
      subprocess.call(('open', OUTPUT_FILE))

if __name__ == "__main__":
  sql_to_xlsx()